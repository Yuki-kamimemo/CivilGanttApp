import webview
import os
import sys
import json
import openpyxl
import re # ★追加：テキストやHTMLタグを処理する機能

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ★追加：画面の備考欄（HTML）の<br>や<div>を、Excel用の改行に変換して綺麗にする関数
def strip_html_tags(text):
    if not text:
        return ""
    # <br>や<div>を改行（\n）に変換
    text = re.sub(r'<(br|div|p)[^>]*>', '\n', text, flags=re.IGNORECASE)
    # その他の不要なタグをすべて削除
    text = re.sub(r'<[^>]+>', '', text)
    # 余分な改行を整理
    text = re.sub(r'\n+', '\n', text)
    return text.strip()

class Api:
    def __init__(self):
        self._window = None
        self._current_file_path = None 
        self._initial_file_content = None # ★追加: 起動時に読み込むデータ
        
    # ★追加: JavaScriptから初期データを受け取るための関数
    def get_initial_data(self):
        return self._initial_file_content

    def save_file(self, data_str, default_filename):
        if not self._window:
            return False
        
        file_types = ('Civil Schedule Master Files (*.csm)', 'All files (*.*)')
        result = self._window.create_file_dialog(
            webview.SAVE_DIALOG, 
            directory='', 
            save_filename=default_filename, 
            file_types=file_types
        )
        
        if result and len(result) > 0:
            file_path = result[0]
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(data_str)
                self._current_file_path = file_path
                return True
            except Exception as e:
                print(f"ファイルの保存に失敗しました: {e}")
                return False
        return False

    def open_file(self):
        if not self._window:
            return None
        
        file_types = ('Civil Schedule Master Files (*.csm)', 'All files (*.*)')
        result = self._window.create_file_dialog(
            webview.OPEN_DIALOG, 
            allow_multiple=False, 
            file_types=file_types
        )
        
        if result and len(result) > 0:
            file_path = result[0]
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                self._current_file_path = file_path
                return content
            except Exception as e:
                print(f"ファイルの読み込みに失敗しました: {e}")
                return None
        return None

    def overwrite_file(self, data_str):
        if not self._current_file_path:
            return False 
            
        try:
            with open(self._current_file_path, 'w', encoding='utf-8') as f:
                f.write(data_str)
            return True
        except Exception as e:
            print(f"上書き保存に失敗しました: {e}")
            return False

    def clear_file_path(self):
        self._current_file_path = None
        return True

    # ★改修：Excelファイルとして本格的な工程表をエクスポートする機能
    def export_to_excel(self, data_str, default_filename):
        if not self._window:
            return False
            
        try:
            # 画面から受け取ったデータを読み解く
            export_data = json.loads(data_str)
            state = export_data.get("state", {})
            national_holidays = export_data.get("nationalHolidays", {})
        except Exception as e:
            print(f"データの読み込みに失敗しました: {e}")
            return False

        # 新しいExcelファイルを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工程表"
        
        # --- Excel用のデザイン（スタイル）設定 ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        
        # --- プロジェクト情報の書き込み ---
        project_name = state.get("projectName", "")
        company_name = state.get("companyName", "")
        p_start = state.get("projectStart", "")
        p_end = state.get("projectEnd", "")
        
        ws.cell(row=1, column=1, value="工 程 表").font = Font(size=18, bold=True)
        ws.cell(row=2, column=1, value=f"工事名: {project_name}")
        ws.cell(row=2, column=3, value=f"事業者名: {company_name}")
        ws.cell(row=2, column=5, value=f"全体工期: {p_start} ～ {p_end}")

        # ★追加：当初予定と変更工程の凡例（注釈）
        ws.cell(row=3, column=5, value="■ 当初予定").font = Font(color="3B82F6", bold=True)
        ws.cell(row=3, column=11, value="■ 変更工程").font = Font(color="DC3545", bold=True)

        # 左側の表の列幅を設定
        ws.column_dimensions['A'].width = 5   # No.
        ws.column_dimensions['B'].width = 15  # 工種
        ws.column_dimensions['C'].width = 15  # 種別
        ws.column_dimensions['D'].width = 20  # 細別・規格
        
        # --- カレンダーの期間計算 ---
        start_date_str = state.get("displayStart") or p_start
        end_date_str = state.get("displayEnd") or p_end
        if not start_date_str or not end_date_str:
            return False
            
        d_start = datetime.strptime(start_date_str, "%Y-%m-%d")
        d_end = datetime.strptime(end_date_str, "%Y-%m-%d")
        total_days = (d_end - d_start).days + 1
        date_list = [d_start + timedelta(days=i) for i in range(total_days)]
        
        # --- ヘッダー（表の見出し）の書き込み ---
        headers = ["No.", "工種", "種別", "細別・規格"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.alignment = align_center
            ws.cell(row=4, column=col).border = thin_border
            ws.cell(row=5, column=col).border = thin_border
            ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col) # 縦に結合
            
        cal_start_col = 5 # カレンダーはE列(5列目)からスタート
        notes_col = cal_start_col + total_days # ★追加：カレンダーのさらに右端を「全体備考」の列にする
        
        # カレンダー列の幅を狭くする
        for i in range(total_days):
            col_letter = get_column_letter(cal_start_col + i)
            ws.column_dimensions[col_letter].width = 3
            
        # ★追加：全体備考のヘッダーを作成
        ws.column_dimensions[get_column_letter(notes_col)].width = 35 # 備考欄の幅
        notes_header = ws.cell(row=4, column=notes_col, value="備考欄")
        notes_header.alignment = align_center
        ws.cell(row=4, column=notes_col).border = thin_border
        ws.cell(row=5, column=notes_col).border = thin_border
        ws.merge_cells(start_row=4, start_column=notes_col, end_row=5, end_column=notes_col)
            
        # --- カレンダーの日付と休日の書き込み ---
        current_month = -1
        month_start_col = cal_start_col
        
        holiday_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid') # 薄いグレー
        national_holiday_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # 薄い赤
        holidays_settings = state.get("holidays", {})
        custom_holidays = holidays_settings.get("custom", [])

        def is_holiday(dt):
            date_str = dt.strftime("%Y-%m-%d")
            if holidays_settings.get("nationalHolidays") and date_str in national_holidays: return "national"
            if holidays_settings.get("sundays") and dt.weekday() == 6: return "holiday"
            if holidays_settings.get("saturdays") and dt.weekday() == 5: return "holiday"
            if date_str in custom_holidays: return "holiday"
            return None

        holiday_cols = {}
        
        for i, dt in enumerate(date_list):
            c_col = cal_start_col + i
            cell_d = ws.cell(row=5, column=c_col, value=dt.day)
            cell_d.alignment = align_center
            cell_d.border = thin_border
            
            if dt.month != current_month:
                if current_month != -1:
                    ws.merge_cells(start_row=4, start_column=month_start_col, end_row=4, end_column=c_col-1)
                month_start_col = c_col
                current_month = dt.month
                cell_m = ws.cell(row=4, column=month_start_col, value=f"{dt.year}年{dt.month}月")
                cell_m.alignment = align_center
                
            ws.cell(row=4, column=c_col).border = thin_border
            
            ht = is_holiday(dt)
            if ht:
                holiday_cols[c_col] = ht
                if ht == "national":
                    cell_d.fill = national_holiday_fill
                    cell_d.font = Font(color="DC3545")
                else:
                    cell_d.fill = holiday_fill
                    cell_d.font = Font(color="6C757D")

        ws.merge_cells(start_row=4, start_column=month_start_col, end_row=4, end_column=cal_start_col + total_days - 1)

        # --- 作業タスク（行）とバーの書き込み ---
        current_row = 6
        tasks = state.get("tasks", [])
        
        for task in tasks:
            periods = task.get("periods", [])
            max_drow = 0
            for p in periods:
                drow = p.get("displayRow", 0)
                if drow > max_drow:
                    max_drow = drow
            
            row_span = max_drow + 1
            end_row = current_row + row_span - 1
            
            ws.cell(row=current_row, column=1, value=task.get("no", "")).alignment = align_center
            ws.cell(row=current_row, column=2, value=task.get("koshu", "")).alignment = align_center
            ws.cell(row=current_row, column=3, value=task.get("shubetsu", "")).alignment = align_center
            ws.cell(row=current_row, column=4, value=task.get("saibetsu", "")).alignment = align_center
            
            for c in range(1, 5):
                for r in range(current_row, end_row + 1):
                    ws.cell(row=r, column=c).border = thin_border
                if row_span > 1:
                    ws.merge_cells(start_row=current_row, start_column=c, end_row=end_row, end_column=c)

            for r in range(current_row, end_row + 1):
                for i in range(total_days):
                    c_col = cal_start_col + i
                    cell = ws.cell(row=r, column=c_col)
                    cell.border = thin_border
                    ht = holiday_cols.get(c_col)
                    if ht == "national":
                        cell.fill = PatternFill(start_color='FFF1F2', end_color='FFF1F2', fill_type='solid')
                    elif ht == "holiday":
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                # ★追加：各行の全体備考列にも枠線を引く
                ws.cell(row=r, column=notes_col).border = thin_border
            
            for p in periods:
                start_str = p.get("start")
                end_str = p.get("end")
                if not start_str or not end_str:
                    continue
                    
                drow = p.get("displayRow", 0)
                color_code = p.get("color", "#3b82f6").replace("#", "")
                bar_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
                
                try:
                    p_s = datetime.strptime(start_str, "%Y-%m-%d")
                    p_e = datetime.strptime(end_str, "%Y-%m-%d")
                except:
                    continue
                    
                for i, dt in enumerate(date_list):
                    if p_s <= dt <= p_e:
                        c_col = cal_start_col + i
                        cell = ws.cell(row=current_row + drow, column=c_col)
                        cell.fill = bar_fill
                        
            current_row += row_span

        # --- ★追加：右側の「全体備考」を結合して書き込む ---
        if current_row > 6:
            notes_html = state.get("notes", "")
            notes_text = strip_html_tags(notes_html) # タグを消して綺麗な文字にする
            notes_cell = ws.cell(row=6, column=notes_col, value=notes_text)
            notes_cell.alignment = Alignment(vertical='top', wrap_text=True) # 上寄せで折り返し
            ws.merge_cells(start_row=6, start_column=notes_col, end_row=current_row - 1, end_column=notes_col)

        # --- ★追加：下部の「日別備考」を書き込む ---
        daily_tabs = state.get("dailyNoteTabs", [])
        daily_data = state.get("dailyNotesData", {})
        
        for tab in daily_tabs:
            tab_id = tab.get("id")
            tab_name = tab.get("name")
            tab_notes = daily_data.get(tab_id, {})
            
            # 左側にタブ名を書き込む
            ws.cell(row=current_row, column=1, value=tab_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for c in range(1, 5):
                ws.cell(row=current_row, column=c).border = thin_border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            
            max_lines = 1 # セルの高さを決めるための改行カウント
            
            # カレンダーの日付ごとに備考を書き込む
            for i, dt in enumerate(date_list):
                c_col = cal_start_col + i
                date_str = dt.strftime("%Y-%m-%d")
                month_str = f"{dt.year}-{dt.month:02d}"
                
                # 日別のデータを探し、なければ月別のデータを探す
                note_html = tab_notes.get(date_str, "")
                if not note_html and dt.day == 1:
                    note_html = tab_notes.get(month_str, "")
                    
                note_text = strip_html_tags(note_html)
                lines = note_text.count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
                    
                cell = ws.cell(row=current_row, column=c_col, value=note_text)
                cell.border = thin_border
                # アプリの画面に合わせて縦書き（textRotation=255）にする
                cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True, textRotation=255)
            
            # 右側の全体備考列の空セルにも枠線を引く
            ws.cell(row=current_row, column=notes_col).border = thin_border
            
            # 行の高さを文字量に合わせて広げる
            ws.row_dimensions[current_row].height = max_lines * 15 + 40
            current_row += 1


        # --- Excelファイルの保存 ---
        file_types = ('Excel Files (*.xlsx)', 'All files (*.*)')
        result = self._window.create_file_dialog(
            webview.SAVE_DIALOG, 
            directory='', 
            save_filename=default_filename, 
            file_types=file_types
        )
        
        if result and len(result) > 0:
            file_path = result[0]
            try:
                wb.save(file_path)
                return True
            except Exception as e:
                print(f"Excelの保存に失敗しました: {e}")
                return False
        return False

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def start_app():
    api = Api()
    # ★追加: ダブルクリックなどでファイルが渡された場合の処理
    if len(sys.argv) > 1:
        initial_file = sys.argv[1]
        if os.path.exists(initial_file):
            try:
                with open(initial_file, 'r', encoding='utf-8') as f:
                    api._initial_file_content = f.read()
                    api._current_file_path = initial_file
            except Exception as e:
                print(f"初期ファイルの読み込みに失敗しました: {e}")
    html_path = get_resource_path('index.html')

    window = webview.create_window(
        'Civil Schedule Master', 
        url=html_path, 
        width=1000, 
        height=700,
        js_api=api
    )
    api._window = window 
    webview.start()

if __name__ == '__main__':
    start_app()