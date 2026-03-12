import json
import re
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ★追加：画面の備考欄（HTML）の<br>や<div>を、Excel用の改行に変換して綺麗にする関数
def strip_html_tags(text):
    if not text:
        return ""
    # <br>や<div>を改行（\n）に変換
    text = re.sub(r'<(br|div|p)[^>]*>', '\n', text, flags=re.IGNORECASE)
    # その他の不要なタグをすべて削除
    text = re.sub(r'<[^>]+>', '', text)
    # 余分な改行を整理
    text = re.sub(r'\n+', '\n', text)
    return text.strip()

def export_data_to_excel(data_str, file_path):
    try:
        export_data = json.loads(data_str)
        state = export_data.get("state", {})
        national_holidays = export_data.get("nationalHolidays", {})
    except Exception as e:
        raise ValueError(f"JSONデータのパースに失敗しました: {e}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # デフォルトの空シートを削除
    
    daily_tabs = state.get("dailyNoteTabs", [])
    if not daily_tabs:
        daily_tabs = [{"id": "tab_general", "name": "工程表"}]
    daily_data = state.get("dailyNotesData", {})

    for tab_info in daily_tabs:
        tab_id = tab_info.get("id")
        tab_name = tab_info.get("name", "Sheet")
        safe_title = re.sub(r'[\\*?:/\[\]]', '', tab_name)[:31]
        ws = wb.create_sheet(title=safe_title)
        
        # 【A3印刷設定】
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # 縦方向は自動（収まらない場合は次ページへ）
        
        # 余白設定 (単位はインチ: 1mm ≒ 0.039インチ)
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = True
        
        # 改ページプレビューモードにする
        ws.sheet_view.view = 'pageBreakPreview'
        
        thin_border = Border(left=Side(style='thin', color='DEE2E6'), right=Side(style='thin', color='DEE2E6'), top=Side(style='thin', color='DEE2E6'), bottom=Side(style='thin', color='DEE2E6'))
        header_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # アプリに近いフォント設定
        font_main = Font(name='Meiryo', size=9)
        font_header = Font(name='Meiryo', size=10, bold=True)
        font_title = Font(name='Meiryo', size=18, bold=True)
        
        project_name = state.get("projectName", "")
        company_name = state.get("companyName", "")
        p_start = state.get("projectStart", "")
        p_end = state.get("projectEnd", "")
        
        # タイトルと工事情報
        ws.row_dimensions[1].height = 37.5 # 50px (1px ≒ 0.75pt)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        title_cell = ws.cell(row=1, column=1, value="工 程 表")
        title_cell.font = font_title
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[2].height = 20
        ws.cell(row=2, column=1, value=f"工事名: {project_name}").font = font_header
        ws.cell(row=2, column=5, value=f"事業者名: {company_name}").font = font_header
        # 全体工期をR2 (18列目) に配置
        ws.cell(row=2, column=18, value=f"全体工期: {p_start} ～ {p_end}").font = font_header
        
        ws.column_dimensions['A'].width = 4   # No.
        ws.column_dimensions['B'].width = 18  # 工種
        ws.column_dimensions['C'].width = 18  # 種別
        ws.column_dimensions['D'].width = 22  # 細別・規格
        
        start_date_str = state.get("displayStart") or p_start
        end_date_str = state.get("displayEnd") or p_end
        if not start_date_str or not end_date_str:
            continue
            
        d_start = datetime.strptime(start_date_str, "%Y-%m-%d")
        d_end = datetime.strptime(end_date_str, "%Y-%m-%d")
        total_days = (d_end - d_start).days + 1
        date_list = [d_start + timedelta(days=i) for i in range(total_days)]
        
        headers = ["No.", "工種", "種別", "細別・規格"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = align_center
            ws.cell(row=4, column=col).border = thin_border
            ws.cell(row=5, column=col).border = thin_border
            ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
            
        cal_start_col = 5
        notes_col = cal_start_col + total_days 
        
        # カレンダー列の幅を4ポイントに設定
        for i in range(total_days):
            ws.column_dimensions[get_column_letter(cal_start_col + i)].width = 4
            
        ws.column_dimensions[get_column_letter(notes_col)].width = 30
        notes_header = ws.cell(row=4, column=notes_col, value="備考欄")
        notes_header.font = font_header
        notes_header.fill = header_fill
        notes_header.alignment = align_center
        ws.cell(row=4, column=notes_col).border = thin_border
        ws.cell(row=5, column=notes_col).border = thin_border
        ws.merge_cells(start_row=4, start_column=notes_col, end_row=5, end_column=notes_col)
            
        current_month = -1
        month_start_col = cal_start_col
        
        # 色定義（アプリのCSSに準拠）
        fill_holiday = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid') # 平日以外の非稼働
        fill_national = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # 祝日
        fill_sat = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid') # 土曜
        fill_sun = PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid') # 日曜
        
        holidays_settings = state.get("holidays", {})
        custom_holidays = holidays_settings.get("custom", [])

        def get_holiday_type(dt):
            date_str = dt.strftime("%Y-%m-%d")
            if holidays_settings.get("nationalHolidays") and date_str in national_holidays: return "national"
            if dt.weekday() == 6: return "sun" # 日曜
            if dt.weekday() == 5: return "sat" # 土曜
            if date_str in custom_holidays: return "holiday"
            # 土日が非稼働設定でない場合でも、曜日としての色はつけたい
            return None

        holiday_cols = {}
        for i, dt in enumerate(date_list):
            c_col = cal_start_col + i
            cell_d = ws.cell(row=5, column=c_col, value=dt.day)
            cell_d.alignment = align_center
            cell_d.border = thin_border
            cell_d.font = font_main
            
            if dt.month != current_month:
                if current_month != -1:
                    ws.merge_cells(start_row=4, start_column=month_start_col, end_row=4, end_column=c_col-1)
                month_start_col = c_col
                current_month = dt.month
                cell_m = ws.cell(row=4, column=month_start_col, value=f"{dt.year}年{dt.month}月")
                cell_m.font = font_header
                cell_m.fill = header_fill
                cell_m.alignment = align_center
                
            ws.cell(row=4, column=c_col).border = thin_border
            
            ht = get_holiday_type(dt)
            holiday_cols[c_col] = ht
            if ht == "national":
                cell_d.fill = fill_national
                cell_d.font = Font(name='Meiryo', size=9, color="DC3545", bold=True)
            elif ht == "sun":
                cell_d.fill = fill_sun
                cell_d.font = Font(name='Meiryo', size=9, color="DC3545")
            elif ht == "sat":
                cell_d.fill = fill_sat
                cell_d.font = Font(name='Meiryo', size=9, color="0D6EFD")
            elif ht == "holiday":
                cell_d.fill = fill_holiday
                cell_d.font = Font(name='Meiryo', size=9, color="6C757D")

        ws.merge_cells(start_row=4, start_column=month_start_col, end_row=4, end_column=cal_start_col + total_days - 1)

        current_row = 6
        
        # 結合状態を管理するための変数
        koshu_start_row = 6
        shubetsu_start_row = 6

        for idx, task in enumerate(state.get("tasks", [])):
            periods = task.get("periods", [])
            max_drow = 0
            for p in periods:
                if p.get("displayRow", 0) > max_drow:
                    max_drow = p.get("displayRow", 0)
            
            row_span = max_drow + 1
            end_row = current_row + row_span - 1
            
            # --- 列1: No. ---
            ws.cell(row=current_row, column=1, value=task.get("no", "")).alignment = align_center
            if row_span > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=end_row, end_column=1)

            # --- 列2: 工種 (タスク跨ぎ結合対応) ---
            is_merge_k = task.get("mergeAboveKoshu", False)
            if idx == 0 or not is_merge_k:
                # 前のブロックがあれば結合を確定
                if current_row > koshu_start_row:
                    ws.merge_cells(start_row=koshu_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                koshu_start_row = current_row
                ws.cell(row=current_row, column=2, value=strip_html_tags(task.get("koshu", ""))).alignment = Alignment(vertical='center', wrap_text=True)
            
            # --- 列3: 種別 (タスク跨ぎ結合対応) ---
            is_merge_s = task.get("mergeAboveShubetsu", False)
            if idx == 0 or not is_merge_s:
                # 前のブロックがあれば結合を確定
                if current_row > shubetsu_start_row:
                    ws.merge_cells(start_row=shubetsu_start_row, start_column=3, end_row=current_row - 1, end_column=3)
                shubetsu_start_row = current_row
                ws.cell(row=current_row, column=3, value=strip_html_tags(task.get("shubetsu", ""))).alignment = Alignment(vertical='center', wrap_text=True)

            # --- 列4: 細別・規格 ---
            ws.cell(row=current_row, column=4, value=strip_html_tags(task.get("saibetsu", ""))).alignment = Alignment(vertical='center', wrap_text=True)
            if row_span > 1:
                ws.merge_cells(start_row=current_row, start_column=4, end_row=end_row, end_column=4)
            
            # 共通設定（枠線・フォント）
            for c in range(1, 5):
                for r in range(current_row, end_row + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin_border
                    cell.font = font_main

            # カレンダーエリアの処理
            for r in range(current_row, end_row + 1):
                ws.row_dimensions[r].height = 18
                for i in range(total_days):
                    c_col = cal_start_col + i
                    cell = ws.cell(row=r, column=c_col)
                    cell.border = thin_border
                    ht = holiday_cols.get(c_col)
                    if ht == "national": cell.fill = PatternFill(start_color='FFF1F2', end_color='FFF1F2', fill_type='solid')
                    elif ht == "sun": cell.fill = PatternFill(start_color='FFFAFA', end_color='FFFAFA', fill_type='solid')
                    elif ht == "sat": cell.fill = PatternFill(start_color='F8FBFF', end_color='F8FBFF', fill_type='solid')
                    elif ht == "holiday": cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                ws.cell(row=r, column=notes_col).border = thin_border
            
            for p in periods:
                start_str = p.get("start")
                end_str = p.get("end")
                if not start_str or not end_str:
                    continue
                    
                drow = p.get("displayRow", 0)
                progress = p.get("progress", 0)
                
                color_code = p.get("color", "#3b82f6").replace("#", "")
                bar_fill_full = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
                if color_code.lower() in ["3b82f6", "0d6efd"]: # 青系
                    bar_fill_light = PatternFill(start_color='C6D9FF', end_color='C6D9FF', fill_type='solid')
                elif color_code.lower() in ["dc3545", "ef4444"]: # 赤系
                    bar_fill_light = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                else:
                    bar_fill_light = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
                
                try:
                    p_s = datetime.strptime(start_str, "%Y-%m-%d")
                    p_e = datetime.strptime(end_str, "%Y-%m-%d")
                except ValueError:
                    continue
                
                duration = (p_e - p_s).days + 1
                progress_days = int(duration * (progress / 100.0))
                
                day_count = 0
                for i, dt in enumerate(date_list):
                    if p_s <= dt <= p_e:
                        day_count += 1
                        c_col = cal_start_col + i
                        cell = ws.cell(row=current_row + drow, column=c_col)
                        if day_count <= progress_days:
                            cell.fill = bar_fill_full
                        else:
                            cell.fill = bar_fill_light
                            
            current_row += row_span

        # 最後に工種・種別の結合を確定（全タスク終了後）
        if current_row > koshu_start_row:
            ws.merge_cells(start_row=koshu_start_row, start_column=2, end_row=current_row - 1, end_column=2)
        if current_row > shubetsu_start_row:
            ws.merge_cells(start_row=shubetsu_start_row, start_column=3, end_row=current_row - 1, end_column=3)

        if current_row > 6:
            notes_text = strip_html_tags(state.get("notes", ""))
            notes_cell = ws.cell(row=6, column=notes_col, value=notes_text)
            notes_cell.alignment = Alignment(vertical='top', wrap_text=True)
            ws.merge_cells(start_row=6, start_column=notes_col, end_row=current_row - 1, end_column=notes_col)

        tab_notes = daily_data.get(tab_id, {})
        tab_merges = state.get("dailyNotesMerges", {}).get(tab_id, {})
        
        ws.cell(row=current_row, column=1, value=tab_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for c in range(1, 5):
            ws.cell(row=current_row, column=c).border = thin_border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        
        # 日報備考欄の高さをご指定の150ポイントに固定
        ws.row_dimensions[current_row].height = 150
        
        skip_count = 0
        for i, dt in enumerate(date_list):
            if skip_count > 0:
                skip_count -= 1
                continue
                
            c_col = cal_start_col + i
            date_str = dt.strftime("%Y-%m-%d")
            month_str = f"{dt.year}-{dt.month:02d}"
            
            colspan = tab_merges.get(date_str, 1)
            
            note_html = tab_notes.get(date_str, "")
            if not note_html and dt.day == 1:
                note_html = tab_notes.get(month_str, "")
                
            note_text = strip_html_tags(note_html)
                
            cell = ws.cell(row=current_row, column=c_col, value=note_text)
            cell.border = thin_border
            # 縦書き（255はExcelの縦書き指定）かつ上揃え、中央揃え
            cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True, textRotation=255)
            
            if colspan > 1:
                # 結合範囲がカレンダーの末尾を超えないように調整
                actual_colspan = min(colspan, total_days - i)
                ws.merge_cells(start_row=current_row, start_column=c_col, end_row=current_row, end_column=c_col + actual_colspan - 1)
                
                # 結合されたセルの枠線を引く
                for c_offset in range(actual_colspan):
                    ws.cell(row=current_row, column=c_col + c_offset).border = thin_border
                    
                skip_count = actual_colspan - 1
        
        ws.cell(row=current_row, column=notes_col).border = thin_border

    try:
        wb.save(file_path)
    except Exception as e:
        raise RuntimeError(f"Excelの保存に失敗しました: {e}")
